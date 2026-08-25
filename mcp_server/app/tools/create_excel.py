"""Tool：產生 Excel（.xlsx）檔案。

用結構化參數（rows/公式字串/圖表定義/樣式），不是讓模型自己寫程式碼跑：
- 公式：儲存格內容以 "=" 開頭的字串會被 openpyxl 原樣寫入，Excel 開啟時自己算，
  這裡不執行任何運算。
- 圖表：openpyxl.chart 支援長條/折線/圓餅，用資料範圍字串驅動。
這樣「進階功能」（公式、圖表、樣式）都能在固定 schema 內做到，不需要 sandbox。

參數用 Pydantic model（不是裸 dict/list[dict]）宣告巢狀欄位：MCP 工具的 JSON
schema 只會照函式簽名自動產生，不會讀函式的 docstring——如果只宣告
`sheets: list[dict]`，模型看到的 schema 只知道「一個物件的陣列」，猜不出物件
裡該放什麼欄位，容易亂猜、驗證失敗、重試到 LangGraph 的 recursion limit
（實測發生過）。用 Pydantic model 才能讓每個巢狀欄位的名稱/型別/說明都出現在
schema 裡，模型才填得對。
"""
import base64
import io
import os

import httpx
from mcp_types import CallToolResult, TextContent
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font
from pydantic import BaseModel, Field

_CHART_TYPES = {"bar": BarChart, "line": LineChart, "pie": PieChart}

# 上傳/下載都是同一個服務（devops.avc.co:18082），跟 personal.js 的
# fileBase 是同一個位址；上傳走「個人」端點，依工號存進
# workspace/<工號>/upload_file/，跟使用者自己上傳附件共用同一套機制。
DEVOPS_FILE_BASE = os.environ.get("DEVOPS_FILE_BASE", "https://devops.avc.co:18082")
DEVOPS_FILE_VERIFY_SSL = os.environ.get("DEVOPS_FILE_VERIFY_SSL", "false").lower() != "false"


def upload_to_workspace(filename: str, data: bytes, mime: str, workid: str) -> str:
    """把產生好的檔案上傳到該工號的 workspace，回傳可以直接用的下載連結。

    workspace_path 回來的是 openclaw 自己看到的路徑（/root/.openclaw/workspace/...），
    但下載服務認的是它自己容器內的路徑（/app/openclaw/workspace/...）——同一份儲存、
    不同容器掛的路徑前綴不同，所以下載連結要用後者，不能直接照抄 workspace_path。
    這裡的字串替換已經拿真實環境測過一輪（上傳→下載）確認可行，不是猜的。
    """
    resp = httpx.post(
        f"{DEVOPS_FILE_BASE}/files/upload/personal",
        params={"workid": workid},
        files={"file": (filename, data, mime)},
        verify=DEVOPS_FILE_VERIFY_SSL,
        timeout=30,
    )
    resp.raise_for_status()
    workspace_path = resp.json()["workspace_path"]
    download_dir = workspace_path.replace("/root/.openclaw/workspace", "/app/openclaw/workspace", 1)
    download_dir = download_dir.rsplit("/", 1)[0]
    from urllib.parse import quote
    return f"{DEVOPS_FILE_BASE}/download?filename={quote(filename)}&download_dir={quote(download_dir)}"


class ChartSpec(BaseModel):
    type: str = Field("bar", description="圖表類型：bar（長條）、line（折線）、pie（圓餅）")
    title: str = Field("", description="圖表標題")
    data_range: str = Field(..., description="資料範圍（不含工作表名），例如 'B2:B5'")
    categories_range: str | None = Field(None, description="類別（X 軸標籤）範圍，例如 'A2:A5'")
    anchor: str = Field("E2", description="圖表放置的儲存格位置，例如 'D2'")
    data_has_header: bool = Field(False, description="data_range 第一格是否為標題（欄位名稱）")


class SheetSpec(BaseModel):
    name: str = Field("Sheet1", description="工作表名稱")
    rows: list[list[str | int | float]] = Field(
        default_factory=list,
        description="每一列的儲存格值，由左到右排列（第一列通常是欄位標題）。"
                    "儲存格內容若以 '=' 開頭會被當成 Excel 公式（例如 '=SUM(B2:B3)'），"
                    "Excel 開啟時自己計算，不是這裡先算好結果。",
    )
    header_bold: bool = Field(False, description="是否把第一列（標題列）文字加粗")
    column_widths: dict[str, float] = Field(
        default_factory=dict, description="欄寬設定，key 是欄位字母（例如 'A'），value 是寬度")
    charts: list[ChartSpec] = Field(default_factory=list, description="要加入這個工作表的圖表清單")


def _add_chart(ws, spec: ChartSpec) -> None:
    chart_cls = _CHART_TYPES.get(spec.type, BarChart)
    chart = chart_cls()
    chart.title = spec.title
    data_ref = Reference(ws, range_string=f"{ws.title}!{spec.data_range}")
    chart.add_data(data_ref, titles_from_data=spec.data_has_header)
    if spec.categories_range:
        chart.set_categories(Reference(ws, range_string=f"{ws.title}!{spec.categories_range}"))
    ws.add_chart(chart, spec.anchor)


def _build(sheets: list[SheetSpec]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for spec in sheets:
        ws = wb.create_sheet(spec.name[:31])
        for r, row in enumerate(spec.rows, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        if spec.header_bold and spec.rows:
            for c in range(1, len(spec.rows[0]) + 1):
                ws.cell(row=1, column=c).font = Font(bold=True)
        for col, width in spec.column_widths.items():
            ws.column_dimensions[col].width = width
        for chart_spec in spec.charts:
            _add_chart(ws, chart_spec)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def register(server) -> None:
    @server.tool(
        name="create_excel",
        description="產生 Excel（.xlsx）檔案，支援多工作表、公式（儲存格內容以 = 開頭）、"
                    "基本圖表（長條/折線/圓餅）與欄寬/粗體樣式。"
                    "何時使用：使用者要求把資料整理成 Excel、報表、清單、對照表。",
    )
    def create_excel(filename: str, sheets: list[SheetSpec]) -> CallToolResult:
        name = filename if filename.endswith(".xlsx") else filename + ".xlsx"
        data = _build(sheets)
        return CallToolResult(
            content=[TextContent(type="text", text=f"已產生 Excel 檔案：{name}")],
            structured_content={
                "filename": name,
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "data_base64": base64.b64encode(data).decode(),
            },
        )

    @server.tool(
        name="create_excel_plain",
        description="產生 Excel（.xlsx）檔案並存進使用者自己的 workspace，回傳可直接下載的連結。"
                    "支援多工作表、公式（儲存格內容以 = 開頭）、基本圖表（長條/折線/圓餅）與欄寬/粗體樣式。"
                    "何時使用：使用者要求把資料整理成 Excel、報表、清單、對照表。"
                    "workid 必填：目前對話代表的員工工號，決定檔案存進哪個人的 workspace。",
    )
    def create_excel_plain(filename: str, sheets: list[SheetSpec], workid: str) -> CallToolResult:
        """給 openclaw 這類不支援 content_and_artifact 分工的 MCP client 用（跟
        knowledge_search_plain 同樣的理由——openclaw 只要偵測到 structured_content
        存在，就會整個蓋掉 content，模型看到的只剩一段 base64，看不到真正有用的
        下載連結）。這裡完全不回 structured_content，真正把檔案存到磁碟（原本的
        create_excel 只存進記憶體、從沒寫過檔案，回傳的下載連結是模型自己編的、
        一定打不開），下載連結由實際上傳結果組出來，不是猜的。
        """
        name = filename if filename.endswith(".xlsx") else filename + ".xlsx"
        data = _build(sheets)
        try:
            url = upload_to_workspace(
                name, data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                workid,
            )
        except Exception as e:  # noqa: BLE001
            return CallToolResult(
                content=[TextContent(type="text", text=f"Excel 已產生但上傳失敗，無法提供下載連結：{e}")],
            )
        return CallToolResult(
            content=[TextContent(type="text", text=f"已產生 Excel 檔案：{name}\n下載連結：{url}")],
        )
